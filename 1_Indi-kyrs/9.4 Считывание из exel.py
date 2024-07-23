# считывание данных из exls файлов

import openpyxl # установим пакет   openpyxl  через команду pip install openpyxl

# откроем нужный файл в режиме только чтения
book = openpyxl.load_workbook(r'E:\док\Энциклопедия эффективности FoE.xlsx', read_only=True)
# read_only=True параметр указывает только на чтение
print(type(book),' - тип нашей книги')
print(book.sheetnames,' - список наименования листов в книге')
sheet = book.active # метод по умолчанию берёт первый лист

# для получения всех листов нужно
sheets = book.worksheets

#для получения конкретного листа
sheet2 = book.worksheets[1]
print(sheet2['A1'].value)

# для получения значения ячейки нужно
print(sheet['A2'].value) # лист[внутренний адрес].значение
print(sheet[2][0].value) # лист[индекс страницы][индекс столбца(начинаются с нуля)].значение
print('______')

print('как перебрать значения в своём диапазоне')
cells = sheet['A24':'b29'] # диапазон ячеек сохраняем в переменную
for name, score in cells:   # переберём эти ячейки с сохранением в свою переменную
    print(name.value, score.value)
print('______')

print('другой способ перебора ячеек')
for row in sheet.iter_rows(min_row=1, max_row=10, min_col=1,max_col=5):  # так мы получаем кортежи
    for cell in row: # переберём кортеж из которого достанем все ячейки
        print(cell.value, end=' ')
    print() #  после вывода одной строки поставим абзац


